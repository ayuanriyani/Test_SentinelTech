import mysql.connector
import base64
import mimetypes

import  decimal
import  math
import  numpy as np
import openpyxl
from urllib.parse import urlparse
from urllib.parse import parse_qs
from datetime import date
import datetime
import pymongo
import json
import os
from dotenv import load_dotenv
load_dotenv()
import requests
from requests.structures import CaseInsensitiveDict
import operator as op
import psycopg2
from psycopg2 import sql, OperationalError


def pystringify(json_str):
    json_str = json.dumps(json_str, separators=(',', ':'))
    json_str = json_str.replace('"', '\\"')
    json_str = json_str.replace('/', '\\\/')
    return json_str

def jsonfy(json_str):
    json_str = json.dumps(json_str, separators=(',', ':'))
    return json_str

def Tupel(setring):
    res = tuple(map(str, setring.split(',')))
    return res

def DownloadFile(URL,save_as):
    # URL = "https://storage.cloud.google.com/dev-genesis/prog-comm/progressive_report_1662966009187324_1.xlsx"
    response = requests.get(URL)
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filename1 = os.path.join(fileDir, 'Temp/', save_as)
    open(filename1, "wb").write(response.content)

def GetMessageFromErrorResponseApi(URL,PARAMS,token):
    headers = CaseInsensitiveDict()
    headers["Accept"] = "application/json"
    headers["Authorization"] = "Bearer " + token
    r = requests.get(url = URL, params = PARAMS, headers = headers)
    data = r.json()
    messages = data['message']
    return messages

def GetAllFromErrorResponseApi(URL,PARAMS,token):
    headers = CaseInsensitiveDict()
    headers["Accept"] = "application/json"
    headers["Authorization"] = "Bearer " + token
    r = requests.get(url = URL, params = PARAMS, headers = headers)
    data = r.json()
    # messages = data['message']
    return data

def GetMessageFromErrorResponseApiPOST(URL,myobject,token):
    headers = CaseInsensitiveDict()
    headers["Accept"] = "application/json"
    headers["Authorization"] = "Bearer " + token
    r = requests.post(URL, json = myobject, headers = headers)
    data = r.json()
    messages = data['message']
    return messages

def GetMessageFromErrorResponseApiPATCH(URL,myobject,token):
    headers = CaseInsensitiveDict()
    headers["Accept"] = "application/json"
    headers["Authorization"] = "Bearer " + token
    r = requests.patch(URL, json = myobject, headers = headers)
    data = r.json()
    return data

def GetSecretDB(env,db_service):
    if env == 'dev' and (db_service == 'horde' or db_service == 'hydra' or db_service == 'gober'):
        host = os.environ.get('mydevhordehost')
        user = os.environ.get('mydevhordeuser')
        password = os.environ.get('mydevhordepass')
        port = 3306
        dbAccess = [host, user, password, port]
    elif env == 'stg' and db_service == 'horde':
        host = os.environ.get('mystghordehost')
        user = os.environ.get('mystghordeuser')
        password = os.environ.get('mystghordepass')
        port = os.environ.get('mystghordeport')
        dbAccess = [host, user, password, port]
    elif env == 'stg' and db_service == 'hydra':
        host = os.environ.get('mystghydrahost')
        user = os.environ.get('mystghydrauser')
        password = os.environ.get('mystghydrapass')
        port = os.environ.get('mystghydraport')
        dbAccess = [host, user, password, port]
    elif env == 'stg' and db_service == 'gober':
        host = os.environ.get('mystggoberhost')
        user = os.environ.get('mystggoberuser')
        password = os.environ.get('mystggoberpass')
        port = os.environ.get('mystggoberport')
        dbAccess = [host, user, password, port]
    elif env == 'prd' and db_service == 'horde':
        host = os.environ.get('myprdhordehost')
        user = os.environ.get('myprdhordeuser')
        password = os.environ.get('myprdhordepass')
        port = 3306
        dbAccess = [host, user, password, port]
    # elif env == 'prd' and db_service == 'hydra':
    # elif env == 'prd' and db_service == 'gober':
    else:
        dbAccess = ''
    return dbAccess

def SqlUpdateCommand(hostq,userq,passwordq,db,sql,portq):
    mydb = mysql.connector.connect(
    host=hostq,
    user=userq,
    password=passwordq,
    database=db,
    port=portq
    )

    mycursor = mydb.cursor()
    mycursor.execute(sql)
    mydb.commit()
    myresult = str(mycursor.rowcount) + " record(s) affected"
    return myresult

def SqlInsertCommand(hostq,userq,passwordq,db,sql,portq):
    mydb = mysql.connector.connect(
    host=hostq,
    user=userq,
    password=passwordq,
    database=db,
    port=portq
    )
    mycursor = mydb.cursor()
    mycursor.execute(sql)
    mydb.commit()
    myresult = mycursor.fetchall()
    return myresult

def SqlSelectCommand(hostq,userq,passwordq,db,sql,portq):
    mydb = mysql.connector.connect(
    host=hostq,
    user=userq,
    password=passwordq,
    database=db,
    port=portq
    )
    mycursor = mydb.cursor()
    mycursor.execute(sql)
    myresult = mycursor.fetchall()
    return myresult

def CreatePostgresqlConnection(db_name, db_user, db_password, db_host, db_port, query):
    """Create a connection to the PostgreSQL database."""
    try:
        # Establish the connection
        connection = psycopg2.connect(
            dbname=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port
        )
        print('Connection to PostgreSQL database successful')
        # Execute query
        cursor = connection.cursor()
        cursor.execute(query)
        connection.commit()
        # Fetch query result
        results = cursor.fetchall()
        connection.close()
        return results
    except OperationalError as e:
        print(f"Error: {e}")
        return None

def PostgresqlUpdateCommand(db_name, db_user, db_password, db_host, db_port, query):
    try:
        # Establish the connection
        connection = psycopg2.connect(
            dbname=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port
        )
        print('Connection to PostgreSQL database successful')
        # Execute query
        cursor = connection.cursor()
        cursor.execute(query)
        connection.commit()
        results = str(cursor.rowcount) + " record(s) affected"
        connection.close()
        print(results)
        return results
    except OperationalError as e:
        print(f"Error: {e}")
        return None
    
def GetMongoConnection(server,collection):
    dev = os.environ.get('dev')
    stg = os.environ.get('stg')
    prd = os.environ.get('prd')
    devhydra = os.environ.get('devhydra')
    stghydra = os.environ.get('stghydra')
    
    if server == 'dev':
        client = pymongo.MongoClient(dev, serverSelectionTimeoutMS=5000)
        mydb = client["horde-dev"]
    elif server == 'stg':
        client = pymongo.MongoClient(stg, serverSelectionTimeoutMS=5000)
        mydb = client["horde-stg"]
    elif server == 'prd':
        client = pymongo.MongoClient(prd, serverSelectionTimeoutMS=5000)
        mydb = client["horde"]
    elif server == 'devhydra':
        client = pymongo.MongoClient(devhydra, serverSelectionTimeoutMS=5000)
        mydb = client['hydra-dev']
    elif server == 'stghydra':
        client = pymongo.MongoClient(stghydra, serverSelectionTimeoutMS=5000)
        mydb = client["hydra-stg"]
    return mydb

def Readfile(namafile):
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filename1 = os.path.join(fileDir, 'SchemaObject/'+namafile)
    with open(filename1, 'r') as file:
        teks = file.read()
        # print("Reading "+namafile)
    return teks

def Writefile(namafile,isifile):
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filename1 = os.path.join(fileDir, 'SchemaObject/'+namafile)
    with open(filename1, 'w') as file:
        file.write(isifile)
    # print("Writing to "+namafile)

def Appendfile(namafile,isifile):
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filename1 = os.path.join(fileDir, 'SchemaObject/'+namafile)
    with open(filename1, 'a') as file:
        file.write(isifile)
    # print("Writing to "+namafile)

def CreateCommodityCsv(data):
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filename1 = os.path.join(fileDir, 'SchemaObject/bulk_upload_commodity.csv')
    f = open(filename1, "w")
    f.write("commodity_code;commodity_name;description;commodity_group;surcharge_applicable;document_surcharge;hs_code;service_type;min_price;is_dangerous_goods\n")
    f = open(filename1, "a")
    f.write(data)
    f.close()

def CreateCsvNoHeader(data,fileName):
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filename1 = os.path.join(fileDir, fileName)
    # f = open(filename1, "w")
    # f.write("commodity_code;commodity_name;description;commodity_group;surcharge_applicable;document_surcharge;hs_code;service_type;min_price\n")
    f = open(filename1, "a")
    f.write(data)
    f.close()

def CreateCsvNoHeaderOverwrite(data,fileName):
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filename1 = os.path.join(fileDir, fileName)
    f = open(filename1, "w")
    f.write(data)
    f.close()

def HapusSpasi(data):
    balikan = data.replace(' ', '')
    return balikan

def MongoCount(server,collection):
    mydb = GetMongoConnection(server, collection)
    mycol = mydb[collection]
    results_count = mycol.count_documents({})
    return results_count

def MongoQuery(server,collection,sortby,q1):
    mydb = GetMongoConnection(server,collection)
    mycol = mydb[collection]
    mydocQ = mycol.find(q1).limit(20).sort(sortby,-1)
    return mydocQ

def RoundDown(x):
    x = int(x)
    return x

def DateLessThan(date1,date2):
    return date1 < date2

def PengurangTanggal(tanggalawal,jumlahHari):
    tanggalawal2 = tanggalawal - datetime.timedelta(days=jumlahHari)
    return tanggalawal2

def AddDay(tanggalawal,penambahanhari):
    tanggalawal2 = tanggalawal + datetime.timedelta(days=penambahanhari)
    return tanggalawal2

def AddHour(tanggalawal,penambahanjam):
    tanggalawal2 = tanggalawal + datetime.timedelta(hours=penambahanjam)
    return tanggalawal2

def ReformatDate(tanggal):
    tanggal = str(tanggal)
    tanggal2 = tanggal[:10] + " " + tanggal[11:19]
    return tanggal2

def ReformatDateIgnoreSecond(tanggal):
    tanggal = str(tanggal)
    tanggal2 = tanggal[:10] + " " + tanggal[11:17]
    return tanggal2

def DateToDate(tanggal):
    return tanggal.strftime("%Y-%m-%d")

def DateToTime(tanggal):
    return tanggal.strftime("%H:%M:%S")

def PanjangString(setring):
    setring = str(setring)
    panjang = len(setring)
    return panjang

def TimeToDecimal(JamMenitDetik):
    JamMenitDetik = str(JamMenitDetik)
    cekformat = len(JamMenitDetik)
    if cekformat == 7:
        JamMenit = '0' + JamMenitDetik
    jam = int(JamMenit[:2])
    menit = int(JamMenit[3:5])
    menitToJam = menit / 60
    JamDesimal = jam + menitToJam
    return JamDesimal

def DateToHour(tanggal):
    return tanggal.strftime("%H")

def DateToDay(tanggal):
    cr_date = datetime.datetime.strptime(tanggal, '%Y-%m-%d %H:%M:%S.%f')
    return cr_date.strftime("%A")

def SelisihWaktu(akhir,awal):
    delta = akhir - awal
    return delta

def CombineDateAndTime(tanggal,jam):
    tanggal = str(tanggal)
    tahun = int(tanggal[:4])
    bulan = int(tanggal[5:7])
    tanggal = int(tanggal[8:10])
    jam = str(jam)
    cekjam = len(jam)
    if cekjam == 7:
        jam = '0' + jam
    start_time = datetime.time(int(jam[:2]), int(jam[3:5]), int(jam[6:8]))
    tanggal = datetime.date(tahun, bulan, tanggal)
    datetime1 = datetime.datetime.combine(tanggal, start_time)
    return datetime1

def SelisihJam(akhir,awal):
    akhir = str(akhir)
    awal = str(awal)
    cekAwal = len(awal)
    if cekAwal == 7:
        awal = '0' + awal
    cekAkhir = len(akhir)
    if cekAkhir == 7:
        akhir = '0' + akhir
    start_time = datetime.time(int(awal[:2]), int(awal[3:5]), int(awal[6:8]))
    stop_time = datetime.time(int(akhir[:2]), int(akhir[3:5]), int(akhir[6:8]))
    date = datetime.date(1, 1, 1)
    datetime1 = datetime.datetime.combine(date, start_time)
    datetime2 = datetime.datetime.combine(date, stop_time)
    time_elapsed = datetime2 - datetime1
    return time_elapsed

def TambahJamMenitDetik(StartingTime,penambah):
    penambah = str(penambah)
    cekLen = len(penambah)
    if cekLen == 7:
        penambah = '0' + penambah
    jam = int(penambah[:2])
    menit = int(penambah[3:5])
    detik = int(penambah[6:8])
    hasil = StartingTime + datetime.timedelta(hours=jam)
    hasil = hasil + datetime.timedelta(minutes=menit)
    hasil = hasil + datetime.timedelta(seconds=detik)
    return hasil

def KurangiJamMenitDetik(StartingTime,penambah):
    penambah = str(penambah)
    cekLen = len(penambah)
    if cekLen == 7:
        penambah = '0' + penambah
    jam = int(penambah[:2])
    menit = int(penambah[3:5])
    detik = int(penambah[6:8])
    hasil = StartingTime - datetime.timedelta(hours=jam)
    hasil = hasil - datetime.timedelta(minutes=menit)
    hasil = hasil - datetime.timedelta(seconds=detik)
    return hasil

def TambahDetik(akhir,awal):
    start_time = datetime.time(int(awal[:2]), int(awal[3:5]), int(awal[6:8]))
    time_elapsed = akhir + start_time
    return time_elapsed

def AmbilParamURL(urlString, KeyNya):
    parsed_url = urlparse(urlString)
    captured_value = parse_qs(parsed_url.query)[KeyNya][0]
    return captured_value

def NyieunList(dari, ke, jarak):
    listnya = np.arange(dari, ke, jarak).tolist()
    return listnya

def GabungString(string1, string2):
    gabungan = str(string1) + str(string2)
    return gabungan

def GabungStringSpasi(string1, string2):
    gabungan = string1 + ' ' + string2
    return gabungan

def PotongString(strings,awal,akhir):
    strings = str(strings)
    a = int(awal)
    b = int(akhir)
    string = strings[a:b]
    return string

def Roundup2(angka):
    result = decimal.Decimal(angka).quantize(decimal.Decimal('0.00'), rounding=decimal.ROUND_UP)
    return result

def Roundup4(angka):
    result = decimal.Decimal(angka).quantize(decimal.Decimal('0.0000'), rounding=decimal.ROUND_UP)
    return result

def Ceiling(angka,pembulatan):
    result = int(math.ceil(angka / pembulatan)) * pembulatan
    return result

def UPPERCASE(teks):
    upper = teks.upper()
    return upper

def setup():
    print 
    "setting stuff up"

def calculate(number1, number2, operation):
    arithmetic_function = { '+': op.add, '-': op.sub, '*': op.mul, '/': op.floordiv }
    return arithmetic_function[operation](number1, number2)

def teardown():
    print 
    "tearing stuff down"

def get_excel_headers(file_path: str) -> list:
    """
    Membuka file Excel dan mengembalikan header (nilai dari baris pertama) sebagai sebuah list.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
            
        return headers
    except Exception as e:
        raise AssertionError(f"Gagal membaca file Excel: {e}")

def validate_column_contains_message(
    file_path: str,
    column_name: str,
    expected_message: str
):
    """
    Memvalidasi bahwa setiap baris pada kolom tertentu mengandung pesan tertentu.
    Baris kosong akan di-skip.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        raise AssertionError(f"Gagal membuka file Excel: {e}")

    headers = [cell.value for cell in sheet[1]]
    if column_name not in headers:
        raise AssertionError(
            f"Kolom '{column_name}' tidak ditemukan. Header tersedia: {headers}"
        )

    target_col_index = headers.index(column_name)
    failures = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        cell_value = row[target_col_index].value

        if not cell_value:
            continue

        cell_value_str = str(cell_value)

        if expected_message not in cell_value_str:
            failures.append(
                f"Baris {row_index}: Tidak mengandung pesan yang diharapkan.\n"
                f"Actual   : '{cell_value_str}'\n"
                f"Expected : '{expected_message}'"
            )

    if failures:
        error_message = (
            f"Validasi kolom '{column_name}' gagal:\n" + "\n".join(failures)
        )
        raise AssertionError(error_message)       

def validate_column_dates_are_within_days(file_path: str, column_name: str, day_limit: int = 7):
    """
    Memvalidasi bahwa semua tanggal di kolom tertentu berada dalam rentang hari yang ditentukan dari hari ini.
    Fungsi ini dapat menangani beberapa format tanggal yang berbeda.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        raise AssertionError(f"Gagal membuka file Excel: {e}")

    headers = [cell.value for cell in sheet[1]]
    if column_name not in headers:
        raise AssertionError(f"Kolom '{column_name}' tidak ditemukan di file. Header yang ada: {headers}")
    target_col_index = headers.index(column_name)

    today_date = datetime.datetime.now(datetime.timezone.utc).date()

    possible_formats = [
        '%Y-%m-%dT%H:%M:%SZ',  # Format asli dengan 'T' dan 'Z'
        '%Y-%m-%d %H:%M',      # Format baru: 2025-08-24 00:00
        '%Y-%m-%d %H:%M:%S'    # Format umum lain (jika ada detik)
    ]

    failures = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        cell_value = row[target_col_index].value
        if not cell_value:
            continue

        parsed_date = None
        cell_value_str = str(cell_value)
        for fmt in possible_formats:
            try:
                parsed_date = datetime.datetime.strptime(cell_value_str, fmt)
                break
            except ValueError:
                continue

        if not parsed_date:
            failures.append(f"Baris {row_index}: Format tanggal tidak valid -> '{cell_value_str}'")
            continue

        if parsed_date.tzinfo is None:
            cell_date_utc = parsed_date.replace(tzinfo=datetime.timezone.utc)
        else:
            cell_date_utc = parsed_date

        cell_date = cell_date_utc.date()
        delta_days = (today_date - cell_date).days

        if delta_days > day_limit:
            failures.append(f"Baris {row_index}: Tanggal '{cell_value_str}' lebih dari {day_limit} hari (selisih {delta_days} hari).")

    if failures:
        error_message = f"Validasi tanggal pada kolom '{column_name}' gagal:\n" + "\n".join(failures)
        raise AssertionError(error_message)
    
def bulk_upload_file_with_form_data(url, token, archive_type, activity_name, file_path, is_scheduler="false"):
    headers = {
        "Authorization": token,
    }
    data = {
        "archive_type": archive_type,
        "activity_name": activity_name,
        "is_scheduler": is_scheduler
    }
    with open(file_path, "rb") as f:
        files = {
            "archive_file": (file_path.split("/")[-1], f, "text/csv")
        }

        s = requests.Session()
        req = requests.Request("POST", url, headers=headers, data=data, files=files)
        prepped = s.prepare_request(req)

        print("\n===== REQUEST HEADERS =====")
        for k, v in prepped.headers.items():
            print(f"{k}: {v}")
        
        print("\n===== REQUEST BODY (short) =====")
        body_preview = prepped.body[:2000] if hasattr(prepped.body, '__getitem__') else str(prepped.body)
        print(body_preview)

        response = s.send(prepped)
        return response.text
def file_to_data_url(file_path, mime_type=None):
    """
    Read an image file and return a data URL: "data:image/jpeg;base64,...".
    Replicates the TypeScript logic provided by the user.
    """
    if not os.path.isabs(file_path):
        file_path = os.path.abspath(file_path)
    
    if mime_type is None:
        mime_type, _ = mimetypes.guess_type(file_path)
    
    if mime_type is None:
        # Fallback for common image extensions if mimetypes fails
        ext = os.path.splitext(file_path)[1].lower()
        mime_map = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp",
        }
        mime_type = mime_map.get(ext)
    
    if not mime_type:
        raise ValueError(
            f"Unknown image mime for {file_path}. Pass mime_type (e.g. 'image/jpeg') "
            f"or use .jpeg, .jpg, .png, .gif, .webp."
        )
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Image file not found: {file_path}")

    with open(file_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
    
    return f"data:{mime_type};base64,{encoded_string}"

def files_to_data_urls(file_paths, mime_type=None):
    """
    Read one or more image files and return a list of data URLs.
    """
    if isinstance(file_paths, str):
        file_paths = [file_paths]
    
    return [file_to_data_url(fp, mime_type) for fp in file_paths]
